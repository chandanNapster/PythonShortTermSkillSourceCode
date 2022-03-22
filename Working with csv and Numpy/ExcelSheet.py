from openpyxl import load_workbook

excel_file = "D:\\UPES\Semester 1 2022\\Online Python Training\\PythonDevWork\\Dataset\\excel.xlsx"


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print("Title of the sheet is = {}".format(sheet.title))
        for row in sheet.rows:
            for cell in row:
                print("The column in EXCEL sheet is: {}{} = The value is: {}".format(
                    cell.column_letter, cell.row, cell.value))


if __name__ == "__main__":
    read_all_data(excel_file)
    # read_all_data("\\Dataset\\excel.xlsx")
