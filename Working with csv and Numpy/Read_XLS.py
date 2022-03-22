from openpyxl.chart import BubbleChart, Reference, Series
from openpyxl.chart import BarChart, LineChart, Reference
import openpyxl
import xlrd

# COMMAND TO CHECK INSTALLED PYTHON PACKAGES
# pip list
# pip install xlrd==1.2.0

workbook = xlrd.open_workbook('excel.xlsx')
excel_sheet = workbook.sheet_by_index(0)

rows = excel_sheet.nrows
cols = excel_sheet.ncols

rowList = []
list = []

for row in range(rows):
    for col in range(cols):
        rowList.append(excel_sheet.cell_value(row, col))

print(rowList)


# for i in range(rows):
#     print(excel_sheet.cell_value(i, 0), excel_sheet.cell_value(
#         i, 1), excel_sheet.cell_value(i, 2))

# excel_sheet.cell_value(0, 0)

# print(excel_sheet.ncols)

# r1 = range(excel_sheet.ncols)

# print(r1)
# print(excel_sheet.nrows)

# for col in range(cols):
#     for row in range(rows):
#         print(excel_sheet.cell_value(row, col))


# DATA WITH OPENPYXL
wb_obj = openpyxl.load_workbook('excel.xlsx')
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)

print(cell_obj)

row = sheet_obj.max_row
column = sheet_obj.max_column

# USE THE NESTED FOR LOOPS TO OUTPUT ALL ELEMENTS STORED IN THE EXCEL FILE

for rw in range(1, row + 1):
    for col in range(1, column + 1):
        cell_obj = sheet_obj.cell(row=rw, column=col)
        print(cell_obj.value)

# READING ALL THE DATA FROM EXCEL SHEET BASED ON EXCEL SHEET COLUMNS
cell_obj = sheet_obj['A1':'C7']

for cell1, cell2, cell3 in cell_obj:
    print(cell1.value, cell2.value, cell3.value)


